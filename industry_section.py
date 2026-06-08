# pyright: reportMissingImports=false
"""
industry_section.py — 네이버 업종별 종목 탐색 섹션
업종 검색(타이핑·초성) → 시가총액순 종목 테이블(종목코드·52주 최저/최고 포함)
→ 선택 종목을 StockReportSection에 전달, 엑셀(.xlsx) 저장 지원
"""
import csv
import json
import threading
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

NAVER_INDUSTRY_LIST   = "https://m.stock.naver.com/api/stocks/industry?page=1&pageSize=100"
NAVER_INDUSTRY_STOCKS = (
    "https://m.stock.naver.com/api/stocks/industry/{no}"
    "?page=1&pageSize=100&sortType=marketValue&sortOrder=desc"
)
NAVER_INTEGRATION = "https://m.stock.naver.com/api/stock/{code}/integration"

# 엑셀/CSV 공통 헤더
EXPORT_HEADER = ["순위", "종목명", "종목코드", "업종", "시가총액",
                 "종가", "등락률(%)", "52주최저", "52주최고", "거래소"]

_CHOSUNG = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"


def _chosung(text: str) -> str:
    """문자열의 한글 음절을 초성으로 변환(영문·기타는 그대로). 초성 검색용."""
    out = []
    for ch in text:
        code = ord(ch)
        if 0xAC00 <= code <= 0xD7A3:
            out.append(_CHOSUNG[(code - 0xAC00) // 588])
        else:
            out.append(ch)
    return "".join(out)


def _fetch_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode("utf-8"))


def _fetch_52w(code: str) -> tuple:
    """단일 종목의 (52주 최저, 52주 최고) 조회. 실패 시 ('-', '-')."""
    try:
        d = _fetch_json(NAVER_INTEGRATION.format(code=code))
        lo = hi = "-"
        for t in d.get("totalInfos", []):
            c = t.get("code")
            if c == "lowPriceOf52Weeks":
                lo = t.get("value", "-")
            elif c == "highPriceOf52Weeks":
                hi = t.get("value", "-")
        return lo, hi
    except Exception:
        return "-", "-"


def _fmt_mktcap(raw: int) -> str:
    if raw >= 1_000_000_000_000:
        return f"{raw / 1_000_000_000_000:.1f}조"
    if raw >= 100_000_000:
        return f"{raw / 100_000_000:.0f}억"
    return f"{raw:,}"


def _ex_code(s: dict) -> str:
    ex = s.get("stockExchangeType", {})
    return ex.get("code", "KS") if isinstance(ex, dict) else str(ex)


class IndustrySectorSection:
    """업종 탐색 섹션. build(container) 로 UI 구성."""

    def __init__(self, root: tk.Tk, theme: dict, report_section, main_notebook=None, sub_notebook=None):
        self.root           = root
        self.report_section = report_section   # StockReportSection 참조
        self.main_notebook  = main_notebook
        self.sub_notebook   = sub_notebook
        self.bg_card        = theme["bg_card"]
        self.bg_input       = theme["bg_input"]
        self.text_light     = theme["text_light"]
        self.text_dark      = theme.get("text_dark", "#1A1A1A")
        self.text_muted     = theme["text_muted"]
        self.accent         = theme["accent"]
        self.btn_rss        = theme["btn_rss"]
        self.tertiary_hov   = theme.get("tertiary_hov", "#003DD6")
        self.primary        = theme.get("primary", "#1A1A1A")
        self._groups: list  = []        # 전체 업종 그룹
        self._stocks: list  = []        # 현재 업종 종목
        self._current_group = None
        self._all_display: list = []    # "업종명 (N종목)" 표시 문자열
        self._all_raw: list     = []    # 업종 원본 이름(검색용)
        self._display_to_group: dict = {}
        self._load_token    = 0         # 52주 로딩 stale 방지 토큰
        self._hide_job      = None

    # ──────────────────────────────────────────────
    # UI 구성
    # ──────────────────────────────────────────────
    def build(self, container: tk.Frame) -> None:
        frame = tk.LabelFrame(
            container,
            text="  📊 업종별 종목 탐색 (한국)  ",
            font=("맑은 고딕", 10, "bold"),
            bg=self.bg_card, fg=self.text_dark,
            bd=3, relief="solid", padx=10, pady=10,
        )
        frame.pack(fill="both", expand=True, pady=4)
        self._frame = frame

        # ── 1행: 업종 검색창 + 새로고침 + 통계 ──
        top_row = tk.Frame(frame, bg=self.bg_card)
        top_row.pack(fill="x", pady=(0, 6))

        tk.Label(top_row, text="업종:",
                 font=("맑은 고딕", 9, "bold"),
                 bg=self.bg_card, fg=self.text_dark).pack(side="left", padx=(0, 4))

        # 타이핑·초성 검색 가능한 입력창
        self.industry_var = tk.StringVar()
        self.industry_entry = ttk.Entry(
            top_row, textvariable=self.industry_var,
            width=24, font=("맑은 고딕", 9),
        )
        self.industry_entry.pack(side="left", padx=(0, 8))
        self.industry_entry.bind("<KeyRelease>", self._on_industry_type)
        self.industry_entry.bind("<Down>",   self._focus_listbox)
        self.industry_entry.bind("<Return>", lambda e: self._commit_first())
        self.industry_entry.bind("<Escape>", lambda e: self._hide_popup())
        self.industry_entry.bind("<FocusIn>",  lambda e: self._on_industry_type(e))
        self.industry_entry.bind("<FocusOut>", self._schedule_hide)

        self.refresh_btn = tk.Button(
            top_row, text="🔄 업종 목록",
            font=("맑은 고딕", 9), bg=self.primary, fg=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=8, pady=4,
            command=self._load_industries_thread,
        )
        self.refresh_btn.pack(side="left", padx=(0, 12))

        self.stats_lbl = tk.Label(
            top_row, text="업종을 불러오는 중...",
            font=("맑은 고딕", 8), bg=self.bg_card, fg=self.text_muted,
        )
        self.stats_lbl.pack(side="left")

        # 검색 자동완성 팝업(Listbox) — frame 위에 place 로 떠 있음
        self._popup = tk.Frame(frame, bg=self.text_dark, bd=1, relief="solid")
        self.listbox = tk.Listbox(
            self._popup, height=10, width=30, font=("맑은 고딕", 9),
            activestyle="none", exportselection=False,
            bg=self.bg_input, fg=self.text_dark,
            selectbackground=self.accent, selectforeground=self.text_light,
            highlightthickness=0, bd=0,
        )
        self.listbox.pack(fill="both", expand=True)
        self.listbox.bind("<ButtonRelease-1>", lambda e: self._commit_listbox())
        self.listbox.bind("<Return>",          lambda e: self._commit_listbox())
        self.listbox.bind("<Escape>",          lambda e: self._hide_popup())

        # ── 2행: 종목 Treeview ──
        tree_frame = tk.Frame(frame, bg=self.bg_card)
        tree_frame.pack(fill="both", expand=True, pady=(0, 6))

        col_cfg = [
            ("순위",     40, "center"),
            ("종목명",   130, "w"),
            ("종목코드",  70, "center"),
            ("시가총액",  85, "e"),
            ("종가",      80, "e"),
            ("등락률",    72, "e"),
            ("52주최저",  80, "e"),
            ("52주최고",  80, "e"),
            ("거래소",    52, "center"),
        ]
        cols = tuple(c[0] for c in col_cfg)
        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="extended", height=8,
        )
        for col, width, anchor in col_cfg:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor=anchor, stretch=False)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.tag_configure("rise", foreground="#E63B2E")
        self.tree.tag_configure("fall", foreground="#0055FF")
        self.tree.tag_configure("odd",  background="#F7F3EC")

        # ── 3행: 선택·실행 버튼 + 엑셀 저장 ──
        btn_row = tk.Frame(frame, bg=self.bg_card)
        btn_row.pack(fill="x")

        for text, cmd, padx in [
            ("상위 5 선택", self._select_top5,  (0, 6)),
            ("전체 선택",   self._select_all,    (0, 6)),
            ("선택 해제",   self._deselect_all,  (0, 16)),
        ]:
            tk.Button(
                btn_row, text=text,
                font=("맑은 고딕", 9), bg=self.bg_input, fg=self.text_dark,
                bd=1, relief="solid", cursor="hand2", padx=8, pady=5,
                command=cmd,
            ).pack(side="left", padx=padx)

        self.analyze_btn = tk.Button(
            btn_row,
            text="📊  선택 종목 리포트 생성",
            font=("맑은 고딕", 10, "bold"),
            bg=self.btn_rss, fg=self.text_light,
            activebackground=self.tertiary_hov, activeforeground=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=16, pady=6,
            command=self._send_to_report,
        )
        self.analyze_btn.pack(side="left")

        # 오른쪽 맨끝: 엑셀 저장(현재 업종 / 전체 업종)
        self.export_mb = tk.Menubutton(
            btn_row, text="💾 엑셀로 저장 ▾",
            font=("맑은 고딕", 9, "bold"),
            bg=self.primary, fg=self.text_light,
            activebackground=self.tertiary_hov, activeforeground=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=12, pady=6,
        )
        export_menu = tk.Menu(self.export_mb, tearoff=0, font=("맑은 고딕", 9))
        export_menu.add_command(label="현재 업종 종목 저장",
                                command=lambda: self._export("current"))
        export_menu.add_command(label="전체 업종 종목 저장",
                                command=lambda: self._export("all"))
        self.export_mb["menu"] = export_menu
        self.export_mb.pack(side="right")

        # 초기 업종 목록 로드
        self._load_industries_thread()

    # ──────────────────────────────────────────────
    # 업종 목록 로드
    # ──────────────────────────────────────────────
    def _load_industries_thread(self):
        self.refresh_btn.configure(state="disabled", text="로딩 중...")
        threading.Thread(target=self._load_industries, daemon=True).start()

    def _load_industries(self):
        try:
            data   = _fetch_json(NAVER_INDUSTRY_LIST)
            groups = data.get("groups", [])
            self._groups = sorted(groups, key=lambda g: g["name"])
            self.root.after(0, self._init_industry_list)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(
                "업종 목록 오류", f"업종 목록 조회 실패:\n{e}"))
        finally:
            self.root.after(0, lambda: self.refresh_btn.configure(
                state="normal", text="🔄 업종 목록"))

    def _init_industry_list(self):
        self._all_raw     = [g["name"] for g in self._groups]
        self._all_display = [f"{g['name']}  ({g['totalCount']}종목)"
                             for g in self._groups]
        self._display_to_group = dict(zip(self._all_display, self._groups))
        if self._groups:
            self.industry_var.set(self._all_display[0])
            self._load_for_group(self._groups[0])

    # ──────────────────────────────────────────────
    # 업종 검색(타이핑·초성) 자동완성
    # ──────────────────────────────────────────────
    def _matches(self, typed: str) -> list:
        typed = typed.strip()
        if not typed:
            return list(self._all_display)
        q  = typed.lower()
        # 순수 자음(미완성 자모)만 입력된 경우에만 초성 검색 적용.
        # 완성된 음절('증' 등)은 substring 으로만 매칭(결과 과확장 방지).
        has_jamo     = any(0x3131 <= ord(c) <= 0x314E for c in typed)
        has_syllable = any(0xAC00 <= ord(c) <= 0xD7A3 for c in typed)
        chosung_mode = has_jamo and not has_syllable
        qc  = _chosung(typed)
        out = []
        for disp, raw in zip(self._all_display, self._all_raw):
            if q in raw.lower():
                out.append(disp)
            elif chosung_mode and qc in _chosung(raw):
                out.append(disp)
        return out

    def _on_industry_type(self, event):
        # 선택된 항목 그대로면(드롭다운 클릭 직후) 재필터하지 않음
        keysym = getattr(event, "keysym", "")
        if keysym in ("Up", "Down", "Return", "Escape", "Tab",
                      "Left", "Right", "Shift_L", "Shift_R"):
            return
        if not self._all_display:
            return
        matches = self._matches(self.industry_var.get())
        self._show_popup(matches)

    def _show_popup(self, matches: list):
        if self._hide_job:
            self.root.after_cancel(self._hide_job)
            self._hide_job = None
        if not matches:
            self._hide_popup()
            return
        self.listbox.delete(0, "end")
        for m in matches:
            self.listbox.insert("end", m)
        self.listbox.configure(height=min(10, len(matches)))
        # 입력창 바로 아래에 배치
        self._popup.place(in_=self.industry_entry, x=0, rely=1.0, y=2, anchor="nw")
        self._popup.lift()

    def _hide_popup(self):
        self._popup.place_forget()

    def _schedule_hide(self, _event):
        # 리스트박스 클릭이 먼저 처리되도록 약간 지연 후 숨김
        self._hide_job = self.root.after(180, self._hide_popup)

    def _focus_listbox(self, _event):
        if self._popup.winfo_ismapped() and self.listbox.size():
            self.listbox.focus_set()
            self.listbox.selection_clear(0, "end")
            self.listbox.selection_set(0)
            self.listbox.activate(0)
        return "break"

    def _commit_first(self):
        if self.listbox.size():
            self._commit_display(self.listbox.get(0))

    def _commit_listbox(self):
        sel = self.listbox.curselection()
        if sel:
            self._commit_display(self.listbox.get(sel[0]))

    def _commit_display(self, display: str):
        self.industry_var.set(display)
        self._hide_popup()
        self.industry_entry.icursor("end")
        group = self._display_to_group.get(display)
        if group:
            self.industry_entry.focus_set()
            self._load_for_group(group)

    # ──────────────────────────────────────────────
    # 업종 선택 시 종목 로드
    # ──────────────────────────────────────────────
    def _load_for_group(self, group: dict):
        self._current_group = group
        self.stats_lbl.configure(
            text=f"{group['name']} 종목 로딩 중...", fg=self.text_muted)
        threading.Thread(
            target=self._load_sector_stocks, args=(group,), daemon=True
        ).start()

    def _load_sector_stocks(self, group: dict):
        try:
            data   = _fetch_json(NAVER_INDUSTRY_STOCKS.format(no=group["no"]))
            stocks = data.get("stocks", [])
            stocks = sorted(stocks,
                            key=lambda s: int(s.get("marketValueRaw", 0)),
                            reverse=True)
            self._stocks = stocks
            self.root.after(0, lambda g=group: self._populate_tree(g))
        except Exception as e:
            self.root.after(0, lambda: self.stats_lbl.configure(
                text=f"오류: {e}", fg="#E63B2E"))

    def _populate_tree(self, group: dict):
        for row in self.tree.get_children():
            self.tree.delete(row)

        cr    = float(group.get("changeRate", 0))
        sign  = "▲" if cr >= 0 else "▼"
        color = "#E63B2E" if cr >= 0 else "#0055FF"
        self.stats_lbl.configure(
            fg=color,
            text=(
                f"{sign} {abs(cr):.2f}%  |  "
                f"상승 {group['riseCount']}  하락 {group['fallCount']}  "
                f"보합 {group['steadyCount']}  (총 {group['totalCount']}종목)"
            ),
        )

        for i, s in enumerate(self._stocks):
            mkt_raw = int(s.get("marketValueRaw", 0))
            ratio   = float(s.get("fluctuationsRatio", 0))

            tags = []
            if ratio > 0:   tags.append("rise")
            elif ratio < 0: tags.append("fall")
            if i % 2 == 1:  tags.append("odd")

            arrow = "▲" if ratio >= 0 else "▼"
            self.tree.insert(
                "", "end", iid=str(i),
                values=(
                    i + 1,
                    s.get("stockName", ""),
                    s.get("itemCode", ""),
                    _fmt_mktcap(mkt_raw),
                    s.get("closePrice", ""),
                    f"{arrow} {abs(ratio):.2f}%",
                    s.get("_low52", "…"),
                    s.get("_high52", "…"),
                    _ex_code(s),
                ),
                tags=tuple(tags),
            )

        # 52주 최저/최고 백그라운드 동시 로드
        self._load_token += 1
        token = self._load_token
        threading.Thread(target=self._fetch_52w_all,
                         args=(self._stocks, token), daemon=True).start()

    def _fetch_52w_all(self, stocks: list, token: int):
        """현재 표시 종목의 52주 최저/최고를 동시 조회 후 행 갱신."""
        def work(i_s):
            i, s = i_s
            lo, hi = _fetch_52w(s.get("itemCode", ""))
            return i, lo, hi

        targets = [(i, s) for i, s in enumerate(stocks)
                   if "_low52" not in s]
        try:
            with ThreadPoolExecutor(max_workers=12) as ex:
                futs = [ex.submit(work, t) for t in targets]
                for fut in as_completed(futs):
                    if token != self._load_token:
                        return
                    i, lo, hi = fut.result()
                    stocks[i]["_low52"]  = lo
                    stocks[i]["_high52"] = hi
                    self.root.after(0, self._update_52w_row, i, lo, hi, token)
        except Exception:
            pass

    def _update_52w_row(self, i, lo, hi, token):
        if token != self._load_token:
            return
        iid = str(i)
        if self.tree.exists(iid):
            self.tree.set(iid, "52주최저", lo)
            self.tree.set(iid, "52주최고", hi)

    # ──────────────────────────────────────────────
    # 선택 헬퍼
    # ──────────────────────────────────────────────
    def _select_top5(self):
        children = self.tree.get_children()
        self.tree.selection_set(children[:5])

    def _select_all(self):
        self.tree.selection_set(self.tree.get_children())

    def _deselect_all(self):
        self.tree.selection_remove(self.tree.get_children())

    # ──────────────────────────────────────────────
    # 엑셀(.xlsx) / CSV 저장
    # ──────────────────────────────────────────────
    def _export(self, scope: str):
        if scope == "current":
            if not self._stocks or not self._current_group:
                messagebox.showwarning("저장 불가", "먼저 업종을 선택해 주세요.")
                return
            default = f"{self._current_group['name']}_종목.xlsx"
        else:
            if not self._groups:
                messagebox.showwarning("저장 불가", "업종 목록을 먼저 불러와 주세요.")
                return
            default = "전체업종_종목.xlsx"

        path = filedialog.asksaveasfilename(
            title="엑셀로 저장",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel 통합 문서", "*.xlsx"), ("CSV (UTF-8)", "*.csv")],
        )
        if not path:
            return
        self.export_mb.configure(state="disabled")
        threading.Thread(target=self._export_worker,
                         args=(scope, path), daemon=True).start()

    def _export_worker(self, scope: str, path: str):
        try:
            if scope == "current":
                self._attach_52w(self._stocks)
                rows  = self._rows_for_stocks(self._stocks,
                                              self._current_group["name"])
                title = self._current_group["name"]
            else:
                rows = []
                total = len(self._groups)
                for gi, g in enumerate(self._groups):
                    self._set_stats(
                        f"전체 저장 중… 업종 {gi+1}/{total}  ({g['name']})",
                        self.text_muted)
                    try:
                        data   = _fetch_json(
                            NAVER_INDUSTRY_STOCKS.format(no=g["no"]))
                        stocks = data.get("stocks", [])
                    except Exception:
                        continue
                    self._attach_52w(stocks)
                    rows += self._rows_for_stocks(stocks, g["name"])
                title = "전체 업종"

            self._write_file(path, rows)
            self.root.after(0, lambda: messagebox.showinfo(
                "저장 완료",
                f"{title} 종목 {len(rows):,}건을 저장했습니다.\n\n{path}"))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(
                "저장 오류", f"엑셀 저장에 실패했습니다:\n{e}"))
        finally:
            self.root.after(0, lambda: self.export_mb.configure(state="normal"))
            self._set_stats_for_current()

    def _attach_52w(self, stocks: list):
        """52주 값이 없는 종목을 동시 조회하여 채움(저장 시 누락 방지)."""
        targets = [(i, s) for i, s in enumerate(stocks) if "_low52" not in s]
        if not targets:
            return

        def work(i_s):
            i, s = i_s
            return (i, *_fetch_52w(s.get("itemCode", "")))

        with ThreadPoolExecutor(max_workers=12) as ex:
            for i, lo, hi in ex.map(work, targets):
                stocks[i]["_low52"]  = lo
                stocks[i]["_high52"] = hi

    def _rows_for_stocks(self, stocks: list, sector_name: str) -> list:
        rows = []
        for i, s in enumerate(stocks):
            rows.append([
                i + 1,
                s.get("stockName", ""),
                s.get("itemCode", ""),
                sector_name,
                _fmt_mktcap(int(s.get("marketValueRaw", 0))),
                s.get("closePrice", ""),
                s.get("fluctuationsRatio", ""),
                s.get("_low52", "-"),
                s.get("_high52", "-"),
                _ex_code(s),
            ])
        return rows

    def _write_file(self, path: str, rows: list):
        if path.lower().endswith(".csv"):
            self._write_csv(path, rows)
            return
        try:
            self._write_xlsx(path, rows)
        except ImportError:
            # openpyxl 미설치 → CSV 폴백
            alt = path[:-5] + ".csv" if path.lower().endswith(".xlsx") else path + ".csv"
            self._write_csv(alt, rows)

    def _write_csv(self, path: str, rows: list):
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(EXPORT_HEADER)
            w.writerows(rows)

    def _write_xlsx(self, path: str, rows: list):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "업종별 종목"

        ws.append(EXPORT_HEADER)
        for r in rows:
            ws.append(r)

        # 헤더 스타일
        head_fill = PatternFill("solid", fgColor="1A1A1A")
        head_font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
        thin      = Side(style="thin", color="D9D9D9")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill      = head_fill
            cell.font      = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        # 열 너비
        widths = [6, 22, 10, 18, 12, 12, 11, 12, 12, 9]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64 + len(EXPORT_HEADER))}{len(rows) + 1}"
        wb.save(path)

    # ──────────────────────────────────────────────
    # 상태표시 헬퍼(스레드 → UI)
    # ──────────────────────────────────────────────
    def _set_stats(self, text: str, color: str):
        self.root.after(0, lambda: self.stats_lbl.configure(text=text, fg=color))

    def _set_stats_for_current(self):
        if self._current_group:
            self.root.after(0, lambda: self._populate_stats(self._current_group))

    def _populate_stats(self, group: dict):
        cr    = float(group.get("changeRate", 0))
        sign  = "▲" if cr >= 0 else "▼"
        color = "#E63B2E" if cr >= 0 else "#0055FF"
        self.stats_lbl.configure(
            fg=color,
            text=(
                f"{sign} {abs(cr):.2f}%  |  "
                f"상승 {group['riseCount']}  하락 {group['fallCount']}  "
                f"보합 {group['steadyCount']}  (총 {group['totalCount']}종목)"
            ),
        )

    # ──────────────────────────────────────────────
    # 선택 종목 → StockReportSection 연동
    # ──────────────────────────────────────────────
    def _send_to_report(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 없음", "리포트에 포함할 종목을 선택해 주세요.\n(Ctrl+클릭으로 복수 선택 가능)")
            return

        tickers = []
        for iid in selected:
            s    = self._stocks[int(iid)]
            code = s.get("itemCode", "")
            tickers.append(f"{code}.{_ex_code(s)}")

        # StockReportSection 티커 입력란 자동 채우기
        self.report_section.ticker_entry.delete(0, "end")
        self.report_section.ticker_entry.insert(0, ", ".join(tickers))

        # 시장을 한국으로 자동 설정
        self.report_section.market_var.set("🇰🇷 한국")

        # 메인 탭 전환 (Stock Report는 인덱스 2)
        if self.main_notebook:
            self.main_notebook.select(2)

        # 리포트 즉시 생성
        self.report_section.start_report_thread()
